"""Excel parser adapter using openpyxl."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from docunav.models.document import FileType
from docunav.parsers.base import (
    ParserAdapter,
    ParserCapabilities,
    ParserResult,
    ParserWarning,
)

logger = logging.getLogger(__name__)

# Regex to extract cell references from formulas
_CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'?[A-Za-z0-9_ ]+'?)!)?"
    r"(?P<col>\$?[A-Z]{1,3})(?P<row>\$?\d+)"
)
_RANGE_REF_RE = re.compile(
    r"(?:(?P<sheet>'?[A-Za-z0-9_ ]+'?)!)?"
    r"(?P<start_col>\$?[A-Z]{1,3})(?P<start_row>\$?\d+)"
    r":(?P<end_col>\$?[A-Z]{1,3})(?P<end_row>\$?\d+)"
)


class OpenpyxlAdapter(ParserAdapter):
    """Wraps openpyxl for Excel workbook inspection."""

    def name(self) -> str:
        return "openpyxl"

    def version(self) -> str:
        try:
            import openpyxl
            return openpyxl.__version__
        except Exception:
            return "unknown"

    def capabilities(self) -> ParserCapabilities:
        return ParserCapabilities(
            supported_types=[FileType.XLSX],
            extracts_text=True,
            extracts_tables=True,
            extracts_layout=False,
            extracts_formulas=True,
            extracts_images=False,
            supports_ocr=False,
        )

    def parse(self, file_path: Path, file_type: FileType) -> ParserResult:
        import openpyxl

        result = ParserResult(
            parser_name=self.name(),
            parser_version=self.version(),
            file_type=file_type,
        )

        try:
            wb = openpyxl.load_workbook(
                str(file_path), data_only=False, read_only=False
            )
        except Exception as e:
            result.errors.append(f"Failed to open workbook: {e}")
            return result

        # Also load data-only version for calculated values
        try:
            wb_data = openpyxl.load_workbook(
                str(file_path), data_only=True, read_only=True
            )
        except Exception:
            wb_data = None

        result.sheet_count = len(wb.sheetnames)
        result.metadata = {
            "sheet_names": wb.sheetnames,
            "has_macros": file_path.suffix.lower() in (".xlsm", ".xltm"),
        }

        # Named ranges
        for defn in wb.defined_names.definedName:
            try:
                result.named_ranges[defn.name] = str(defn.attr_text)
            except Exception:
                pass

        # External links
        if hasattr(wb, "_external_links"):
            for link in wb._external_links:
                try:
                    result.external_links.append(str(link.file_link.Target))
                except Exception:
                    pass

        if result.external_links:
            result.warnings.append(
                ParserWarning(
                    code="EXTERNAL_LINKS",
                    message=f"Workbook references {len(result.external_links)} external file(s).",
                )
            )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_data = wb_data[sheet_name] if wb_data else None
            sheet_data = self._extract_sheet(ws, ws_data, sheet_name, result)
            result.sheets.append(sheet_data)

        wb.close()
        if wb_data:
            wb_data.close()

        return result

    def _extract_sheet(
        self,
        ws: Any,
        ws_data: Any | None,
        sheet_name: str,
        result: ParserResult,
    ) -> dict[str, Any]:
        from openpyxl.utils import get_column_letter

        is_visible = ws.sheet_state == "visible"
        if not is_visible:
            result.warnings.append(
                ParserWarning(
                    code="HIDDEN_SHEET",
                    message=f"Sheet '{sheet_name}' is hidden.",
                    sheet=sheet_name,
                )
            )

        # Detect hidden rows and columns
        hidden_rows: list[int] = []
        hidden_cols: list[int] = []

        try:
            for row_idx in range(1, ws.max_row + 1 if ws.max_row else 1):
                rd = ws.row_dimensions.get(row_idx)
                if rd and rd.hidden:
                    hidden_rows.append(row_idx)
        except Exception:
            pass

        try:
            for col_idx in range(1, ws.max_column + 1 if ws.max_column else 1):
                col_letter = get_column_letter(col_idx)
                cd = ws.column_dimensions.get(col_letter)
                if cd and cd.hidden:
                    hidden_cols.append(col_idx)
        except Exception:
            pass

        if hidden_rows:
            result.warnings.append(
                ParserWarning(
                    code="HIDDEN_ROWS",
                    message=f"Sheet '{sheet_name}' has {len(hidden_rows)} hidden row(s).",
                    sheet=sheet_name,
                )
            )
        if hidden_cols:
            result.warnings.append(
                ParserWarning(
                    code="HIDDEN_COLUMNS",
                    message=f"Sheet '{sheet_name}' has {len(hidden_cols)} hidden column(s).",
                    sheet=sheet_name,
                )
            )

        # Detect merged cells
        merged = [str(m) for m in ws.merged_cells.ranges]

        # Extract cells, formulas, and comments
        sheet_cells: list[dict[str, Any]] = []
        sheet_formulas: list[dict[str, Any]] = []
        sheet_comments: list[dict[str, str]] = []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if cell.value is None and cell.comment is None:
                    continue

                addr = cell.coordinate
                col_letter = get_column_letter(cell.column)
                is_hidden = cell.row in hidden_rows or cell.column in hidden_cols

                # Get calculated value from data-only workbook
                calc_value = None
                if ws_data:
                    try:
                        calc_value = ws_data[addr].value
                    except Exception:
                        pass

                raw_value = cell.value
                formula_text = None
                if isinstance(raw_value, str) and raw_value.startswith("="):
                    formula_text = raw_value
                    display_value = calc_value if calc_value is not None else raw_value
                else:
                    display_value = raw_value

                cell_data = {
                    "sheet_name": sheet_name,
                    "address": addr,
                    "value": display_value,
                    "formula": formula_text,
                    "data_type": cell.data_type,
                    "number_format": cell.number_format,
                    "is_hidden": is_hidden,
                    "is_merged": any(addr in str(m) for m in ws.merged_cells.ranges),
                    "row": cell.row,
                    "column": cell.column,
                }
                sheet_cells.append(cell_data)
                result.cells.append(cell_data)

                if formula_text:
                    refs = self._parse_formula_refs(formula_text, sheet_name)
                    formula_data = {
                        "sheet_name": sheet_name,
                        "cell_address": addr,
                        "formula_text": formula_text,
                        "calculated_value": calc_value,
                        **refs,
                    }
                    sheet_formulas.append(formula_data)
                    result.formulas.append(formula_data)

                if cell.comment:
                    comment_data = {
                        "cell": addr,
                        "author": cell.comment.author or "",
                        "text": cell.comment.text or "",
                    }
                    sheet_comments.append(comment_data)

        # Build used_range string
        used_range = None
        if max_row > 0 and max_col > 0:
            start_col = get_column_letter(1)
            end_col = get_column_letter(max_col)
            used_range = f"{start_col}1:{end_col}{max_row}"

        return {
            "name": sheet_name,
            "index": wb_index(ws),
            "is_visible": is_visible,
            "used_range": used_range,
            "row_count": max_row,
            "column_count": max_col,
            "cells": sheet_cells,
            "formulas": sheet_formulas,
            "merged_cells": merged,
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_cols,
            "comments": sheet_comments,
            "formula_count": len(sheet_formulas),
        }

    def _parse_formula_refs(
        self, formula: str, current_sheet: str
    ) -> dict[str, Any]:
        referenced_cells: list[str] = []
        referenced_ranges: list[str] = []
        referenced_sheets: set[str] = set()
        external_refs: list[str] = []

        # Find range references first
        for match in _RANGE_REF_RE.finditer(formula):
            sheet = match.group("sheet")
            if sheet:
                sheet = sheet.strip("'")
                referenced_sheets.add(sheet)
            ref = match.group(0)
            referenced_ranges.append(ref)

        # Find cell references
        remaining = _RANGE_REF_RE.sub("", formula)
        for match in _CELL_REF_RE.finditer(remaining):
            sheet = match.group("sheet")
            if sheet:
                sheet = sheet.strip("'")
                referenced_sheets.add(sheet)
            ref = match.group(0)
            referenced_cells.append(ref)

        # Detect external references
        if "[" in formula:
            external_refs.append(formula)

        return {
            "referenced_cells": referenced_cells,
            "referenced_ranges": referenced_ranges,
            "referenced_sheets": list(referenced_sheets),
            "external_references": external_refs,
        }


def wb_index(ws: Any) -> int:
    """Get the index of a worksheet in its parent workbook."""
    try:
        return ws.parent.sheetnames.index(ws.title)
    except Exception:
        return 0
