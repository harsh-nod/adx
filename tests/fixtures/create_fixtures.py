#!/usr/bin/env python3
"""
Generate Excel (.xlsx) test fixture files for the DocuNav project.

Creates:
  - excel/sample_model.xlsx  -- multi-sheet financial model with formulas,
                                named ranges, a hidden sheet, and hidden rows.
  - excel/simple.xlsx        -- minimal workbook with basic tabular data.

Requirements:
  pip install openpyxl

Usage:
  python create_fixtures.py            # generates files in the same fixtures dir
  python create_fixtures.py /out/dir   # generates files under a custom root
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

def _fixtures_root() -> Path:
    """Return the root fixtures directory, honouring an optional CLI arg."""
    if len(sys.argv) > 1:
        return Path(sys.argv[1])
    return Path(__file__).resolve().parent


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CURRENCY_FMT = '#,##0.00'
PERCENT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row: int, max_col: int) -> None:
    """Apply header styling to an entire row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# sample_model.xlsx
# ---------------------------------------------------------------------------

def _create_sample_model(out_dir: Path) -> Path:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "4472C4"

    # Column widths
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 18
    ws_summary.column_dimensions["D"].width = 18
    ws_summary.column_dimensions["E"].width = 18

    # Title
    ws_summary["A1"] = "Financial Summary"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14)

    # Headers (row 3)
    headers = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4"]
    for col_idx, header in enumerate(headers, start=1):
        ws_summary.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws_summary, 3, len(headers))

    # Data rows -- values reference the Assumptions sheet where applicable
    rows = [
        # (label, Y1, Y2, Y3, Y4)  -- formulas reference Assumptions sheet
        ("Revenue",
         "=Assumptions!B2*Assumptions!B3",
         "=Assumptions!C2*Assumptions!C3",
         "=Assumptions!D2*Assumptions!D3",
         "=Assumptions!E2*Assumptions!E3"),
        ("Cost of Goods Sold (COGS)",
         "=B4*Assumptions!B5",
         "=C4*Assumptions!C5",
         "=D4*Assumptions!D5",
         "=E4*Assumptions!E5"),
        ("Gross Profit",
         "=B4-B5",
         "=C4-C5",
         "=D4-D5",
         "=E4-E5"),
        ("Operating Expenses",
         "=Assumptions!B6",
         "=Assumptions!C6",
         "=Assumptions!D6",
         "=Assumptions!E6"),
        ("EBITDA",
         "=B6-B7",
         "=C6-C7",
         "=D6-D7",
         "=E6-E7"),
        ("Depreciation & Amortization",
         "=Assumptions!B7",
         "=Assumptions!C7",
         "=Assumptions!D7",
         "=Assumptions!E7"),
        ("EBIT",
         "=B8-B9",
         "=C8-C9",
         "=D8-D9",
         "=E8-E9"),
        ("Interest Expense",
         "=Assumptions!B8",
         "=Assumptions!C8",
         "=Assumptions!D8",
         "=Assumptions!E8"),
        ("Earnings Before Tax",
         "=B10-B11",
         "=C10-C11",
         "=D10-D11",
         "=E10-E11"),
        ("Tax",
         "=B12*Assumptions!B9",
         "=C12*Assumptions!C9",
         "=D12*Assumptions!D9",
         "=E12*Assumptions!E9"),
        ("Net Income",
         "=B12-B13",
         "=C12-C13",
         "=D12-D13",
         "=E12-E13"),
    ]

    for row_offset, (label, *values) in enumerate(rows):
        row_num = 4 + row_offset
        ws_summary.cell(row=row_num, column=1, value=label).font = Font(
            bold=(label in ("Gross Profit", "EBITDA", "EBIT", "Net Income")),
        )
        for col_idx, val in enumerate(values, start=2):
            cell = ws_summary.cell(row=row_num, column=col_idx, value=val)
            cell.number_format = CURRENCY_FMT
            cell.border = THIN_BORDER

    # Gross-margin % row (row 15)
    ws_summary.cell(row=15, column=1, value="Gross Margin %").font = Font(italic=True)
    for col_idx in range(2, 6):
        col_letter = get_column_letter(col_idx)
        cell = ws_summary.cell(
            row=15,
            column=col_idx,
            value=f"={col_letter}6/{col_letter}4",
        )
        cell.number_format = PERCENT_FMT

    # Net-margin % row (row 16)
    ws_summary.cell(row=16, column=1, value="Net Margin %").font = Font(italic=True)
    for col_idx in range(2, 6):
        col_letter = get_column_letter(col_idx)
        cell = ws_summary.cell(
            row=16,
            column=col_idx,
            value=f"={col_letter}14/{col_letter}4",
        )
        cell.number_format = PERCENT_FMT

    # Hide rows 15-16 (margin rows) to simulate "hidden rows"
    ws_summary.row_dimensions[15].hidden = True
    ws_summary.row_dimensions[16].hidden = True

    # ---- Sheet 2: Assumptions --------------------------------------------
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_assumptions.sheet_properties.tabColor = "70AD47"

    ws_assumptions.column_dimensions["A"].width = 30
    for c in "BCDE":
        ws_assumptions.column_dimensions[c].width = 16

    headers_a = ["Assumption", "Year 1", "Year 2", "Year 3", "Year 4"]
    for col_idx, h in enumerate(headers_a, start=1):
        ws_assumptions.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws_assumptions, 1, len(headers_a))

    assumption_data = [
        # (label, Y1, Y2, Y3, Y4, fmt)
        ("Units Sold",          10000,  12500,  15000,  18000,  '#,##0'),
        ("Price per Unit",      49.99,  52.49,  54.99,  57.49,  CURRENCY_FMT),
        ("Growth Rate",         None,   0.25,   0.20,   0.20,   PERCENT_FMT),
        ("COGS % of Revenue",   0.40,   0.38,   0.36,   0.35,  PERCENT_FMT),
        ("Operating Expenses",  125000, 140000, 155000, 170000, CURRENCY_FMT),
        ("D&A",                 15000,  16500,  18000,  19500,  CURRENCY_FMT),
        ("Interest Expense",    8000,   7500,   7000,   6500,   CURRENCY_FMT),
        ("Tax Rate",            0.25,   0.25,   0.25,   0.25,   PERCENT_FMT),
    ]

    for row_offset, (label, *values_and_fmt) in enumerate(assumption_data):
        row_num = 2 + row_offset
        values = values_and_fmt[:-1]
        fmt = values_and_fmt[-1]
        ws_assumptions.cell(row=row_num, column=1, value=label)
        for col_idx, val in enumerate(values, start=2):
            cell = ws_assumptions.cell(row=row_num, column=col_idx, value=val)
            cell.number_format = fmt
            cell.border = THIN_BORDER

    # ---- Sheet 3: Hidden -------------------------------------------------
    ws_hidden = wb.create_sheet("Hidden")
    ws_hidden.sheet_state = "hidden"

    ws_hidden["A1"] = "Internal Calculation Reference"
    ws_hidden["A1"].font = Font(bold=True)

    ws_hidden["A2"] = "Discount Factor"
    ws_hidden["B2"] = 0.10
    ws_hidden["B2"].number_format = PERCENT_FMT

    ws_hidden["A3"] = "Risk Premium"
    ws_hidden["B3"] = 0.045
    ws_hidden["B3"].number_format = PERCENT_FMT

    ws_hidden["A4"] = "Terminal Growth Rate"
    ws_hidden["B4"] = 0.025
    ws_hidden["B4"].number_format = PERCENT_FMT

    ws_hidden["A5"] = "WACC"
    ws_hidden["B5"] = "=B2+B3"
    ws_hidden["B5"].number_format = PERCENT_FMT

    ws_hidden["A7"] = "Model Version"
    ws_hidden["B7"] = "3.2.1"

    ws_hidden["A8"] = "Last Calibrated"
    ws_hidden["B8"] = "2026-04-15"

    # ---- Named ranges ----------------------------------------------------
    # Revenue_Y1 -> Summary!B4
    dn_revenue = DefinedName("Revenue_Y1", attr_text="Summary!$B$4")
    wb.defined_names.add(dn_revenue)

    # Net_Income_Y1 -> Summary!B14
    dn_net = DefinedName("Net_Income_Y1", attr_text="Summary!$B$14")
    wb.defined_names.add(dn_net)

    # Assumptions_Table -> Assumptions!A1:E9
    dn_assumptions = DefinedName("Assumptions_Table", attr_text="Assumptions!$A$1:$E$9")
    wb.defined_names.add(dn_assumptions)

    # WACC -> Hidden!B5
    dn_wacc = DefinedName("WACC", attr_text="Hidden!$B$5")
    wb.defined_names.add(dn_wacc)

    # ---- Save ------------------------------------------------------------
    dest = out_dir / "sample_model.xlsx"
    wb.save(str(dest))
    return dest


# ---------------------------------------------------------------------------
# simple.xlsx
# ---------------------------------------------------------------------------

def _create_simple(out_dir: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["Product", "Region", "Sales"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))

    data = [
        ("Laptop",   "North", 12450),
        ("Laptop",   "South", 9800),
        ("Monitor",  "North", 7620),
        ("Monitor",  "South", 6340),
        ("Keyboard", "North", 3210),
        ("Keyboard", "South", 2870),
        ("Mouse",    "North", 1540),
        ("Mouse",    "South", 1380),
    ]

    for row_offset, (product, region, sales) in enumerate(data):
        row_num = 2 + row_offset
        ws.cell(row=row_num, column=1, value=product)
        ws.cell(row=row_num, column=2, value=region)
        cell = ws.cell(row=row_num, column=3, value=sales)
        cell.number_format = '#,##0'

    # Auto-fit rough column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12

    dest = out_dir / "simple.xlsx"
    wb.save(str(dest))
    return dest


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    root = _fixtures_root()
    excel_dir = _ensure_dir(root / "excel")

    print(f"Fixtures root : {root}")
    print(f"Excel directory: {excel_dir}")
    print()

    path1 = _create_sample_model(excel_dir)
    print(f"  Created: {path1}")

    path2 = _create_simple(excel_dir)
    print(f"  Created: {path2}")

    print()
    print("Done. All Excel fixtures generated successfully.")


if __name__ == "__main__":
    main()
